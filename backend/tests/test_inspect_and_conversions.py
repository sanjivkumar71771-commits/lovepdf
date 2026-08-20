"""Tests for /api/pdf/inspect new fields + pdf-to-word / pdf-to-excel / health regression."""
import io
import os

import pytest
import requests
from dotenv import dotenv_values
from fpdf import FPDF

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/")
API = f"{BASE_URL}/api/pdf"

LOHIT = "/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf"


def make_english_pdf(pages=1) -> bytes:
    pdf = FPDF()
    for i in range(pages):
        pdf.add_page()
        pdf.set_font("Helvetica", size=14)
        pdf.cell(0, 10, f"This is page {i + 1} of the test document.", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 10, "The quick brown fox jumps over the lazy dog and then runs away.", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 10, "We have a lot of text here for the test with common English words.", new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())


def make_table_pdf() -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    rows = [
        ["Name", "Age", "City"],
        ["Alice", "30", "Delhi"],
        ["Bob", "25", "Mumbai"],
        ["Carol", "41", "Pune"],
    ]
    for r in rows:
        pdf.cell(60, 10, r[0], border=1)
        pdf.cell(30, 10, r[1], border=1)
        pdf.cell(60, 10, r[2], border=1, new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())


def _kruti_named_ttf() -> str:
    """Clone the Lohit TTF but rename its internal name table to 'KrutiDev010'
    so the resulting PDF advertises a legacy Devanagari font name."""
    out = "/tmp/KrutiDev010.ttf"
    if os.path.exists(out):
        return out
    from fontTools.ttLib import TTFont
    f = TTFont(LOHIT)
    name = f["name"]
    for rec in name.names:
        if rec.nameID in (1, 3, 4, 6, 16):
            name.setName("KrutiDev010", rec.nameID, rec.platformID, rec.platEncID, rec.langID)
    f.save(out)
    return out


def make_kruti_pdf() -> bytes:
    pdf = FPDF()
    pdf.add_font("KrutiDev010", "", _kruti_named_ttf())
    pdf.add_page()
    pdf.set_font("KrutiDev010", size=18)
    pdf.cell(0, 12, "pfj= izek.k i=", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 12, "izekf.kr fd;k tkrk gS", new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())


def post_file(path, data: bytes, name="test.pdf", extra=None):
    files = {"file": (name, data, "application/pdf")}
    return requests.post(f"{API}{path}", files=files, data=extra or {}, timeout=180)


# ---------- health ----------
class TestHealth:
    def test_health_all_tools(self):
        r = requests.get(f"{API}/health", timeout=60)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["ok"] is True
        tools = data["tools"]
        for t in ["soffice", "gs", "qpdf", "tesseract", "pdftoppm", "ocrmypdf"]:
            assert tools.get(t) is True, f"{t} missing: {tools}"


# ---------- /inspect ----------
class TestInspect:
    def test_inspect_english_pdf(self):
        r = post_file("/inspect", make_english_pdf(), "english.pdf")
        assert r.status_code == 200, r.text[:300]
        d = r.json()
        for k in ["legacy_hindi", "has_text", "devanagari_ratio", "looks_english", "fonts"]:
            assert k in d, f"missing key {k}: {d}"
        assert d["has_text"] is True, d
        assert d["legacy_hindi"] is False, d
        assert d["looks_english"] is True, d
        assert d["devanagari_ratio"] == pytest.approx(0.0, abs=0.01), d
        assert isinstance(d["fonts"], list)

    def test_inspect_kruti_like_pdf(self):
        r = post_file("/inspect", make_kruti_pdf(), "kruti.pdf")
        assert r.status_code == 200, r.text[:300]
        d = r.json()
        assert d["has_text"] is True, d
        assert d["legacy_hindi"] is True, d
        assert d["looks_english"] is False, d

    def test_inspect_garbage_file_no_500(self):
        r = post_file("/inspect", b"not a pdf at all", "junk.pdf")
        assert r.status_code == 200, r.text[:300]
        d = r.json()
        assert d["has_text"] is False and d["legacy_hindi"] is False


# ---------- conversions ----------
class TestConversions:
    def test_pdf_to_word(self):
        r = post_file("/pdf-to-word", make_english_pdf(2), "doc.pdf")
        assert r.status_code == 200, r.text[:300]
        assert len(r.content) > 1000, len(r.content)
        from docx import Document
        doc = Document(io.BytesIO(r.content))
        text = "\n".join(p.text for p in doc.paragraphs)
        assert "quick brown fox" in text.replace("\n", " ") or len(text.strip()) > 10, text[:300]

    def test_pdf_to_excel(self):
        r = post_file("/pdf-to-excel", make_table_pdf(), "table.pdf")
        assert r.status_code == 200, r.text[:300]
        assert len(r.content) > 500, len(r.content)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        joined = " ".join(cells)
        assert "Alice" in joined or "Name" in joined, joined[:300]
