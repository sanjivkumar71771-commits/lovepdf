"""Backend integration tests for PDFPro Studio tools."""
import io
import os
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://lovepdf-tools-1.preview.emergentagent.com').rstrip('/')
API = f"{BASE_URL}/api/pdf"

TEXT_PDF = "/tmp/text.pdf"
SCAN_PDF = "/tmp/scan.pdf"


def _docx_text(content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs)


def _xlsx_text(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True)
    parts = []
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            for c in row:
                if c: parts.append(str(c))
    return " ".join(parts)


# ---------- Health ----------
def test_health():
    r = requests.get(f"{API}/health", timeout=30)
    assert r.status_code == 200
    j = r.json()
    assert j["ok"] is True
    for k in ["soffice", "gs", "tesseract", "pdftoppm", "ocrmypdf"]:
        assert j["tools"].get(k) is True, f"{k} missing/false"


# ---------- PDF -> Word: text-based ----------
def test_pdf_to_word_text():
    with open(TEXT_PDF, "rb") as f:
        r = requests.post(f"{API}/pdf-to-word", files={"file": ("text.pdf", f, "application/pdf")}, timeout=120)
    assert r.status_code == 200, r.text[:300]
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.wordprocessingml")
    txt = _docx_text(r.content)
    assert len(txt.strip()) > 0, "docx has no text"


# ---------- PDF -> Word: scanned/OCR ----------
def test_pdf_to_word_scanned_ocr():
    with open(SCAN_PDF, "rb") as f:
        r = requests.post(f"{API}/pdf-to-word",
                          files={"file": ("scan.pdf", f, "application/pdf")},
                          data={"lang": "eng"}, timeout=180)
    assert r.status_code == 200, r.text[:300]
    txt = _docx_text(r.content)
    assert len(txt.strip()) > 5, f"OCR docx should contain text, got: {txt!r}"


# ---------- PDF -> Excel scanned ----------
def test_pdf_to_excel_scanned_ocr():
    with open(SCAN_PDF, "rb") as f:
        r = requests.post(f"{API}/pdf-to-excel", files={"file": ("scan.pdf", f, "application/pdf")}, timeout=180)
    assert r.status_code == 200, r.text[:300]
    txt = _xlsx_text(r.content)
    assert len(txt.strip()) > 5, f"xlsx should contain OCR text, got: {txt!r}"


# ---------- Protect ----------
def test_protect_pdf():
    with open(TEXT_PDF, "rb") as f:
        r = requests.post(f"{API}/protect",
                          files={"file": ("text.pdf", f, "application/pdf")},
                          data={"password": "secret123"}, timeout=60)
    assert r.status_code == 200, r.text[:300]
    assert r.content[:4] == b"%PDF"
    # verify encrypted
    import pikepdf
    with pytest.raises(pikepdf.PasswordError):
        pikepdf.open(io.BytesIO(r.content))
    with pikepdf.open(io.BytesIO(r.content), password="secret123") as pdf:
        assert len(pdf.pages) >= 1


# ---------- Repair ----------
def test_repair_pdf():
    with open(TEXT_PDF, "rb") as f:
        r = requests.post(f"{API}/repair", files={"file": ("text.pdf", f, "application/pdf")}, timeout=60)
    assert r.status_code == 200, r.text[:300]
    assert r.content[:4] == b"%PDF"
